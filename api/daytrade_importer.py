"""
デイトレ xlsx パース & MySQL 挿入ロジック
DayTradeDataFilter/main.py の parse_daytrade_sheet_fixed + build_round_trips を
win32com 依存なしでポート。
"""

import io
import os
import re
import unicodedata
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime, date, time as dt_time

import pandas as pd
import mysql.connector


# ---------------------------------------------------------------------------
# DB接続（既存importer.pyと共通）
# ---------------------------------------------------------------------------

def get_db_config() -> dict:
    return {
        "host": "mysql",
        "port": 3306,
        "database": os.environ["MYSQL_DATABASE"],
        "user": os.environ["MYSQL_USER"],
        "password": os.environ["MYSQL_PASSWORD"],
        "charset": "utf8mb4",
    }


@dataclass
class DaytradeImportResult:
    filename: str
    parsed: int
    inserted: int
    skipped: int
    total_pnl: float
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# パーサー（DayTradeDataFilter/main.py より移植）
# ---------------------------------------------------------------------------

_CODE_RE = re.compile(r"(?P<code>\d{3,4}[A-Z]?)")


def _split_brand(blob: str):
    """'サンリオ   8136   東証' -> (銘柄名, コード)"""
    if not isinstance(blob, str):
        return None, None
    norm = unicodedata.normalize("NFKC", blob).replace("\u00A0", " ").replace("\u3000", " ")
    txt = re.sub(r"\s+", " ", norm).strip()
    m = _CODE_RE.search(txt.upper())
    if not m:
        return (txt or None), None
    code = m.group("code")
    name = txt[:m.start()].strip() or None
    return name, code


def _norm(s) -> str:
    if s is None:
        return ""
    return str(s).replace("\u00A0", " ").replace("\u3000", " ").strip()


def parse_daytrade_sheet(content: bytes, sheet_name: str = "元データ") -> pd.DataFrame:
    """xlsx バイト列をパースして正規化 DataFrame を返す。
    指定シートが存在しない場合は最初のシートを使う。
    """
    bio = io.BytesIO(content)
    import openpyxl
    wb = openpyxl.load_workbook(bio, read_only=True)
    actual_sheet = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
    wb.close()
    bio.seek(0)
    df = pd.read_excel(bio, sheet_name=actual_sheet, header=None)
    n = df.shape[0]
    recs = []
    r = 0
    while r < n:
        v = df.iat[r, 0] if 0 <= r < n else None
        if pd.notna(v) and isinstance(v, (int, float)) and float(v).is_integer():
            base = r
            order_no    = int(v)
            status      = df.iat[base, 1]     if pd.notna(df.iat[base, 1])     else None
            order_type  = df.iat[base, 2]     if pd.notna(df.iat[base, 2])     else None
            brand_blob  = df.iat[base, 3]     if pd.notna(df.iat[base, 3])     else None
            brand, code = _split_brand(brand_blob if isinstance(brand_blob, str) else "")

            trade       = df.iat[base+1, 3]   if base+1 < n and pd.notna(df.iat[base+1, 3]) else None
            order_qty   = df.iat[base+1, 6]   if base+1 < n and pd.notna(df.iat[base+1, 6]) else None
            exec_qty    = int(order_qty) if order_qty is not None and not pd.isna(order_qty) else None
            exec_cond   = df.iat[base+1, 7]   if base+1 < n and pd.notna(df.iat[base+1, 7]) else None
            order_price = df.iat[base+1, 8]   if base+1 < n and pd.notna(df.iat[base+1, 8]) else None

            order_type_str = str(order_type or "")
            offset = 1 if "逆指" in order_type_str else 0
            exec_row = base + 3 + offset
            time_row = base + 4 + offset

            exec_price   = df.iat[exec_row, 7] if exec_row < n and pd.notna(df.iat[exec_row, 7]) else None
            excel_serial = df.iat[exec_row, 5] if exec_row < n and pd.notna(df.iat[exec_row, 5]) else None
            tm           = df.iat[time_row, 5] if time_row < n and pd.notna(df.iat[time_row, 5]) else None

            exec_dt = None
            try:
                if isinstance(excel_serial, (int, float)):
                    exec_date = pd.Timestamp("1899-12-30") + pd.Timedelta(days=float(excel_serial))
                else:
                    exec_date = pd.to_datetime(excel_serial) if excel_serial is not None else None

                if isinstance(tm, dt_time):
                    t = tm
                elif isinstance(tm, str):
                    t = pd.to_datetime(tm).time()
                elif isinstance(tm, (datetime, pd.Timestamp)):
                    t = tm.time() if isinstance(tm, datetime) else tm.to_pydatetime().time()
                else:
                    t = None

                if exec_date is not None and t is not None:
                    exec_dt = pd.Timestamp.combine(exec_date.date(), t)
            except Exception:
                exec_dt = None

            recs.append({
                "注文番号":   order_no,
                "注文種別":   order_type,
                "銘柄名":     brand,
                "銘柄コード": code,
                "取引":       trade,
                "注文株数":   order_qty,
                "執行条件":   exec_cond,
                "注文単価":   order_price,
                "約定株数":   exec_qty,
                "約定単価":   exec_price,
                "約定日時":   exec_dt,
                "_注文状況":  status,
                "_base_row":  base + 1,
            })
            r = base + 1
            continue
        r += 1

    out = pd.DataFrame(recs)
    if not out.empty:
        for col in ["注文株数", "注文単価", "約定株数", "約定単価"]:
            if col in out.columns:
                out[col] = pd.to_numeric(out[col], errors="coerce")
        if "約定日時" in out.columns:
            out["約定日時"] = pd.to_datetime(out["約定日時"], errors="coerce")
    return out


def drop_unwanted(df: pd.DataFrame) -> pd.DataFrame:
    """取消完了 / 現物買 / 現物売 を除外。"""
    if df.empty:
        return df
    mask_cancel = pd.Series(False, index=df.index)
    if "_注文状況" in df.columns:
        mask_cancel = df["_注文状況"].astype(str).map(_norm).eq("取消完了")
    mask_genbutsu = pd.Series(False, index=df.index)
    if "取引" in df.columns:
        mask_genbutsu = df["取引"].astype(str).map(_norm).isin(["現物買", "現物売"])
    return df.loc[~(mask_cancel | mask_genbutsu)].reset_index(drop=True)


def build_round_trips(df_norm: pd.DataFrame) -> list[dict]:
    """エントリー/エグジットをペアリングしてラウンドトリップのリストを返す。"""
    req = ["銘柄コード", "銘柄名", "取引", "約定日時", "約定単価", "約定株数", "注文番号"]
    lacks = [c for c in req if c not in df_norm.columns]
    if lacks:
        raise ValueError(f"必要列が不足: {lacks}")

    df = df_norm.sort_values(["約定日時", "注文番号"], na_position="last").reset_index(drop=True)

    def _side(t):
        t = "" if t is None else str(t)
        if "信新買" in t: return "LONG"
        if "信新売" in t: return "SHORT"
        return None

    def _is_exit(side, t):
        t = "" if t is None else str(t)
        return (side == "LONG" and "信返売" in t) or (side == "SHORT" and "信返買" in t)

    def _num(x):
        try:
            return None if x is None or pd.isna(x) else float(x)
        except Exception:
            return None

    queues = defaultdict(lambda: {"LONG": deque(), "SHORT": deque()})
    trips = []

    for _, row in df.iterrows():
        code  = row["銘柄コード"]
        name  = row["銘柄名"]
        trade = row["取引"]
        ts    = row["約定日時"]
        exec_px = _num(row.get("約定単価"))
        qty     = row.get("約定株数")
        on      = row.get("注文番号")

        if code is None or qty is None or pd.isna(qty) or exec_px is None or pd.isna(ts):
            continue
        qty = int(qty)
        side = _side(trade)

        if side in ("LONG", "SHORT"):
            queues[code][side].append({
                "ts": ts, "px": exec_px, "qty_remaining": qty,
            })
            continue

        for closing_side in ("LONG", "SHORT"):
            if not _is_exit(closing_side, trade):
                continue
            remain = qty
            agg_qty, agg_sum = 0, 0.0
            entry_time_first = None

            while remain > 0 and queues[code][closing_side]:
                ent = queues[code][closing_side][0]
                use = min(remain, ent["qty_remaining"])
                remain -= use
                ent["qty_remaining"] -= use
                if entry_time_first is None:
                    entry_time_first = ent["ts"]
                agg_qty += use
                agg_sum += ent["px"] * use
                if ent["qty_remaining"] == 0:
                    queues[code][closing_side].popleft()

            if agg_qty > 0:
                entry_avg = agg_sum / agg_qty
                exit_px   = exec_px
                pnl = (exit_px - entry_avg) * agg_qty if closing_side == "LONG" else (entry_avg - exit_px) * agg_qty
                dur = int((ts - entry_time_first).total_seconds()) if pd.notna(entry_time_first) and pd.notna(ts) else None
                trips.append({
                    "trade_date":   entry_time_first.date() if pd.notna(entry_time_first) else None,
                    "entry_at":     entry_time_first.to_pydatetime() if pd.notna(entry_time_first) else None,
                    "exit_at":      ts.to_pydatetime() if pd.notna(ts) else None,
                    "duration_sec": dur,
                    "stock_code":   code,
                    "stock_name":   name,
                    "side":         closing_side,
                    "quantity":     agg_qty,
                    "entry_price":  round(entry_avg, 2),
                    "exit_price":   round(exit_px, 2),
                    "pnl":          round(pnl, 2),
                })
    return trips


# ---------------------------------------------------------------------------
# MySQL 挿入
# ---------------------------------------------------------------------------

_INSERT_SQL = """
    INSERT IGNORE INTO daytrades
        (trade_date, entry_at, exit_at, duration_sec, stock_code, stock_name,
         side, quantity, entry_price, exit_price, pnl, source_file, imported_at)
    VALUES
        (%(trade_date)s, %(entry_at)s, %(exit_at)s, %(duration_sec)s,
         %(stock_code)s, %(stock_name)s, %(side)s, %(quantity)s,
         %(entry_price)s, %(exit_price)s, %(pnl)s, %(source_file)s, %(imported_at)s)
"""


def run_daytrade_import(filename: str, content: bytes) -> DaytradeImportResult:
    # 1. パース
    try:
        df_raw = parse_daytrade_sheet(content)
    except Exception as e:
        # シート名が異なる場合などにフォールバック
        raise ValueError(f"xlsx のパースに失敗しました: {e}")

    df_clean = drop_unwanted(df_raw)
    trips = build_round_trips(df_clean)

    if not trips:
        return DaytradeImportResult(
            filename=filename, parsed=0, inserted=0, skipped=0, total_pnl=0.0,
            warnings=["ラウンドトリップが0件でした（元データシートを確認してください）"],
        )

    # 2. DB挿入
    now = datetime.now()
    for t in trips:
        t["source_file"] = filename
        t["imported_at"] = now

    conn = mysql.connector.connect(**get_db_config())
    cursor = conn.cursor()
    try:
        cursor.executemany(_INSERT_SQL, trips)
        inserted = cursor.rowcount
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()

    total_pnl = sum(t["pnl"] for t in trips)
    return DaytradeImportResult(
        filename=filename,
        parsed=len(trips),
        inserted=inserted,
        skipped=len(trips) - inserted,
        total_pnl=round(total_pnl, 2),
    )
